"""
Label Production Web App
========================
Upload Excel files → download generated Word label documents.
Deploy to Render or run locally with: python app.py
"""

import os
import io
import zipfile
import tempfile
from flask import Flask, render_template, request, send_file, flash, redirect, url_for
from lxml import etree
import openpyxl

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "label-production-2026")

# ── XML Namespaces ──────────────────────────────────────────────────────────
WNS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
W14 = "http://schemas.microsoft.com/office/word/2010/wordml"
R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

LABELS_PER_PAGE = 10

CONGREGATE_LAYOUT = [
    ("default", 36, "empty"),
    ("center",  36, "field:School"),
    ("center",  40, "field:Menu_Item"),
    ("center",  28, "prefix_field:# of Servings :Value"),
    ("default", 28, "empty"),
    ("center",  24, "empty"),
    ("center",  24, "field:Meal_Type"),
    ("default", 24, "empty"),
    ("default", 24, "empty"),
    ("default", 24, "empty"),
]

BULK_LAYOUT = [
    ("center",  32, "empty"),
    ("center",  32, "field:School"),
    ("center",  36, "field:Menu_Item"),
    ("center",  32, "prefix_field:# of Servings :Value"),
    ("default", 32, "empty"),
    ("default", 18, "empty"),
    ("left",    18, "two_fields:Type:Delivery"),
    ("center",  18, "empty"),
    ("default", 18, "empty"),
]

BREAKDOWN_LAYOUT = [
    ("center",  32, "empty"),
    ("center",  32, "field:School"),
    ("center",  36, "field:Menu_Item"),
    ("center",  32, "field:Room"),
    ("center",  32, "prefix_field:# of servings :Value"),
    ("default", 32, "empty"),
    ("center",  32, "field:Delivery_"),
]


# ── XML Builder ─────────────────────────────────────────────────────────────

def make_elem(tag, attrib=None, text=None):
    e = etree.Element(f"{{{WNS}}}{tag}")
    if attrib:
        for k, v in attrib.items():
            if ":" in k:
                prefix, local = k.split(":", 1)
                ns = WNS if prefix == "w" else W14
                e.set(f"{{{ns}}}{local}", str(v))
            else:
                e.set(f"{{{WNS}}}{k}", str(v))
    if text is not None:
        e.text = text
    return e


def make_run(text, sz, bold=True):
    r = make_elem("r")
    rpr = make_elem("rPr")
    if bold:
        rpr.append(make_elem("b"))
        rpr.append(make_elem("bCs"))
    rpr.append(make_elem("sz", {"w:val": str(sz)}))
    rpr.append(make_elem("szCs", {"w:val": str(sz)}))
    r.append(rpr)
    t = make_elem("t")
    t.text = str(text)
    if str(text).startswith(" ") or str(text).endswith(" "):
        t.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
    r.append(t)
    return r


def make_paragraph(align, sz, content_type, record):
    p = make_elem("p")
    ppr = make_elem("pPr")
    ppr.append(make_elem("spacing", {"w:before": "111"}))
    ppr.append(make_elem("ind", {"w:left": "115", "w:right": "115"}))
    ppr.append(make_elem("contextualSpacing"))
    if align != "default":
        ppr.append(make_elem("jc", {"w:val": align}))
    rpr = make_elem("rPr")
    rpr.append(make_elem("b"))
    rpr.append(make_elem("bCs"))
    rpr.append(make_elem("sz", {"w:val": str(sz)}))
    rpr.append(make_elem("szCs", {"w:val": str(sz)}))
    ppr.append(rpr)
    p.append(ppr)

    if content_type == "empty":
        pass
    elif content_type.startswith("field:"):
        field_name = content_type.split(":", 1)[1]
        value = record.get(field_name, "")
        if value is None:
            value = ""
        p.append(make_run(str(value), sz))
    elif content_type.startswith("prefix_field:"):
        parts = content_type.split(":", 2)
        prefix, field_name = parts[1], parts[2]
        value = record.get(field_name, "")
        if value is None:
            value = ""
        try:
            num = float(value)
            value = str(int(num)) if num == int(num) else str(num)
        except (ValueError, TypeError):
            value = str(value)
        p.append(make_run(prefix, sz))
        p.append(make_run(value, sz))
    elif content_type.startswith("two_fields:"):
        parts = content_type.split(":", 2)
        v1 = str(record.get(parts[1], "") or "")
        v2 = str(record.get(parts[2], "") or "")
        p.append(make_run(v1, sz))
        tab_r = make_elem("r")
        tab_rpr = make_elem("rPr")
        tab_rpr.append(make_elem("sz", {"w:val": str(sz)}))
        tab_rpr.append(make_elem("szCs", {"w:val": str(sz)}))
        tab_r.append(tab_rpr)
        tab_r.append(make_elem("tab"))
        p.append(tab_r)
        p.append(make_run(v2, sz))
    return p


def make_label_cell(layout, record, width=4320):
    tc = make_elem("tc")
    tcPr = make_elem("tcPr")
    tcPr.append(make_elem("tcW", {"w:w": str(width), "w:type": "dxa"}))
    tcBorders = make_elem("tcBorders")
    for side in ["top", "left", "bottom", "right"]:
        tcBorders.append(make_elem(side, {"w:val": "none", "w:sz": "0", "w:space": "0", "w:color": "auto"}))
    tcPr.append(tcBorders)
    tc.append(tcPr)
    for (a, s, c) in layout:
        tc.append(make_paragraph(a, s, c, record))
    return tc


def make_empty_cell(width):
    tc = make_elem("tc")
    tcPr = make_elem("tcPr")
    tcPr.append(make_elem("tcW", {"w:w": str(width), "w:type": "dxa"}))
    tcBorders = make_elem("tcBorders")
    for side in ["top", "left", "bottom", "right"]:
        tcBorders.append(make_elem(side, {"w:val": "none", "w:sz": "0", "w:space": "0", "w:color": "auto"}))
    tcPr.append(tcBorders)
    tc.append(tcPr)
    p = make_elem("p")
    ppr = make_elem("pPr")
    ppr.append(make_elem("spacing", {"w:before": "0", "w:after": "0"}))
    rpr = make_elem("rPr")
    rpr.append(make_elem("sz", {"w:val": "2"}))
    rpr.append(make_elem("szCs", {"w:val": "2"}))
    ppr.append(rpr)
    p.append(ppr)
    tc.append(p)
    return tc


def make_table(layout, records):
    tbl = make_elem("tbl")
    tblPr = make_elem("tblPr")
    tblPr.append(make_elem("tblW", {"w:w": "9072", "w:type": "dxa"}))
    tblPr.append(make_elem("tblLayout", {"w:type": "fixed"}))
    tblBorders = make_elem("tblBorders")
    for side in ["top", "left", "bottom", "right", "insideH", "insideV"]:
        tblBorders.append(make_elem(side, {"w:val": "none", "w:sz": "0", "w:space": "0", "w:color": "auto"}))
    tblPr.append(tblBorders)
    tbl.append(tblPr)
    tblGrid = make_elem("tblGrid")
    for w in [4320, 432, 4320]:
        tblGrid.append(make_elem("gridCol", {"w:w": str(w)}))
    tbl.append(tblGrid)
    empty_record = {}
    for row_idx in range(5):
        left_idx = row_idx * 2
        right_idx = row_idx * 2 + 1
        left_rec = records[left_idx] if left_idx < len(records) else empty_record
        right_rec = records[right_idx] if right_idx < len(records) else empty_record
        tr = make_elem("tr")
        trPr = make_elem("trPr")
        trPr.append(make_elem("trHeight", {"w:val": "2880", "w:hRule": "exact"}))
        tr.append(trPr)
        tr.append(make_label_cell(layout, left_rec, 4320))
        tr.append(make_empty_cell(432))
        tr.append(make_label_cell(layout, right_rec, 4320))
        tbl.append(tr)
        if row_idx < 4:
            spacer = make_elem("tr")
            sPr = make_elem("trPr")
            sPr.append(make_elem("trHeight", {"w:val": "180", "w:hRule": "exact"}))
            spacer.append(sPr)
            for w in [4320, 432, 4320]:
                spacer.append(make_empty_cell(w))
            tbl.append(spacer)
    return tbl


def create_docx_bytes(layout, records):
    nsmap = {
        "w": WNS, "w14": W14, "r": R_NS,
        "mc": "http://schemas.openxmlformats.org/markup-compatibility/2006",
        "o": "urn:schemas-microsoft-com:office:office",
        "m": "http://schemas.openxmlformats.org/officeDocument/2006/math",
        "v": "urn:schemas-microsoft-com:vml",
        "wp": "http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing",
    }
    doc = etree.Element(f"{{{WNS}}}document", nsmap=nsmap)
    body = etree.SubElement(doc, f"{{{WNS}}}body")
    num_pages = max(1, (len(records) + LABELS_PER_PAGE - 1) // LABELS_PER_PAGE)
    for page_idx in range(num_pages):
        start = page_idx * LABELS_PER_PAGE
        page_records = records[start:start + LABELS_PER_PAGE]
        body.append(make_table(layout, page_records))
        if page_idx < num_pages - 1:
            p = make_elem("p")
            ppr = make_elem("pPr")
            rpr = make_elem("rPr")
            rpr.append(make_elem("sz", {"w:val": "2"}))
            rpr.append(make_elem("szCs", {"w:val": "2"}))
            ppr.append(rpr)
            p.append(ppr)
            r = make_elem("r")
            r.append(make_elem("br", {"w:type": "page"}))
            p.append(r)
            body.append(p)
    sectPr = make_elem("sectPr")
    sectPr.append(make_elem("pgSz", {"w:w": "12240", "w:h": "15840"}))
    sectPr.append(make_elem("pgMar", {
        "w:top": "360", "w:right": "1714", "w:bottom": "0",
        "w:left": "1714", "w:header": "720", "w:footer": "720", "w:gutter": "0",
    }))
    body.append(sectPr)

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"><Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/><Default Extension="xml" ContentType="application/xml"/><Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/></Types>')
        zf.writestr("_rels/.rels", '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/></Relationships>')
        zf.writestr("word/_rels/document.xml.rels", '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"></Relationships>')
        zf.writestr("word/document.xml", etree.tostring(doc, xml_declaration=True, encoding="UTF-8", standalone=True))
    buf.seek(0)
    return buf


# ── Data Reading ────────────────────────────────────────────────────────────

def read_congregate(file_bytes, sheet_name):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[sheet_name]
    records = []
    if sheet_name == "MergeInfo":
        for row in ws.iter_rows(min_row=2, values_only=True):
            school, value, menu_item, meal_type = row[0], row[1], row[2], row[3]
            if school and str(school).strip() and menu_item and str(menu_item).strip():
                records.append({"School": str(school).strip(), "Value": value,
                                "Menu_Item": str(menu_item).strip(),
                                "Meal_Type": str(meal_type).strip() if meal_type else ""})
    else:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 5:
                continue
            school, value, menu_item, meal_type = row[1], row[2], row[3], row[4]
            if school and str(school).strip() and menu_item and str(menu_item).strip():
                records.append({"School": str(school).strip(), "Value": value,
                                "Menu_Item": str(menu_item).strip(),
                                "Meal_Type": str(meal_type).strip() if meal_type else ""})
    wb.close()
    return records


def read_bulk(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb["MergeLabels"]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        school, value, menu_item, delivery, type_ = row[0], row[1], row[2], row[3], row[4]
        if school and str(school).strip() and menu_item and str(menu_item).strip():
            val_str = ""
            if value is not None:
                try:
                    num = float(value)
                    val_str = str(int(num)) if num == int(num) else str(num)
                except (ValueError, TypeError):
                    val_str = str(value)
            records.append({"School": str(school).strip(), "Value": val_str,
                            "Menu_Item": str(menu_item).strip(),
                            "Delivery": str(delivery).strip() if delivery else "",
                            "Type": str(type_).strip() if type_ else ""})
    wb.close()
    return records


def read_breakdown(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb["Sheet3"]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 5:
            continue
        school, room, value, menu_item, delivery = row[0], row[1], row[2], row[3], row[4]
        if school and str(school).strip() and menu_item and str(menu_item).strip():
            records.append({"School": str(school).strip(), "Room": str(room).strip() if room else "",
                            "Value": value, "Menu_Item": str(menu_item).strip(),
                            "Delivery_": str(delivery).strip() if delivery else ""})
    wb.close()
    return records


def get_sheet_names(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    sheets = wb.sheetnames
    wb.close()
    day_order = ["MergeInfo", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "SATSUN"]
    ordered = [s for s in day_order if s in sheets]
    for s in sheets:
        if s not in ordered and s != "Portion":
            ordered.append(s)
    return ordered


# ── Routes ──────────────────────────────────────────────────────────────────

@app.route("/", methods=["GET"])
def index():
    return render_template("index.html")


@app.route("/get-tabs", methods=["POST"])
def get_tabs():
    f = request.files.get("congregate_file")
    if not f:
        return {"tabs": []}
    try:
        tabs = get_sheet_names(f.read())
        return {"tabs": tabs}
    except Exception as e:
        return {"tabs": [], "error": str(e)}


@app.route("/generate", methods=["POST"])
def generate():
    cong_file = request.files.get("congregate_file")
    bulk_file = request.files.get("bulk_file")
    cong_tab = request.form.get("congregate_tab", "MergeInfo")
    date_label = request.form.get("date_label", "").strip() or "Labels"

    if not cong_file and not bulk_file:
        flash("Please upload at least one Excel file.")
        return redirect(url_for("index"))

    results = {}

    try:
        if cong_file:
            cong_bytes = cong_file.read()
            cong_records = read_congregate(cong_bytes, cong_tab)
            if cong_records:
                results["congregate"] = create_docx_bytes(CONGREGATE_LAYOUT, cong_records)

        if bulk_file:
            bulk_bytes = bulk_file.read()
            bulk_records = read_bulk(bulk_bytes)
            if bulk_records:
                results["bulk"] = create_docx_bytes(BULK_LAYOUT, bulk_records)
            breakdown_records = read_breakdown(bulk_bytes)
            if breakdown_records:
                results["breakdown"] = create_docx_bytes(BREAKDOWN_LAYOUT, breakdown_records)
    except Exception as e:
        flash(f"Error processing files: {e}")
        return redirect(url_for("index"))

    if not results:
        flash("No label data found in the uploaded files.")
        return redirect(url_for("index"))

    # Bundle all generated files into a single zip for download
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w") as zf:
        if "congregate" in results:
            zf.writestr(f"Cong {date_label} Labels.docx", results["congregate"].getvalue())
        if "bulk" in results:
            zf.writestr(f"Bulk {date_label} Label.docx", results["bulk"].getvalue())
        if "breakdown" in results:
            zf.writestr(f"Breakdown {date_label} Label.docx", results["breakdown"].getvalue())
    zip_buf.seek(0)

    return send_file(zip_buf, mimetype="application/zip",
                     as_attachment=True, download_name=f"Labels {date_label}.zip")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
